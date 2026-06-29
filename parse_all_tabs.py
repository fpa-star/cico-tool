"""
parse_all_tabs.py
Parses ALL relevant tabs from CICO_ claude.xlsx and updates:
  - data/classification_rules.json   (from Classification Rules tab)
  - data/account_master.json         (from Claude Classification - Section 1)
  - data/valid_heads.json            (from Claude Classification - Section 3)
  - data/pattern_rules_raw.json      (from Claude Classification - Section 2)
  - data/gotcha_rules.json           (from Claude Classification - Section 4)
"""

import json, os, sys
import openpyxl

_HERE  = os.path.dirname(os.path.abspath(__file__))
_DATA  = os.path.join(_HERE, "data")
INPUT  = r"Z:\CICO\CICO_ claude.xlsx"

def _s(v):
    if v is None: return ""
    return str(v).encode("utf-8","replace").decode("utf-8").strip()


# ── 1. Parse Classification Rules tab ─────────────────────────────────────────

def parse_classification_rules(wb):
    ws = wb["Classification Rules"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        vals = [_s(c) for c in row]
        if "Account Name" in vals:
            header_idx = i; break
    if header_idx is None:
        print("  WARNING: Could not find header in Classification Rules"); return []

    header = [_s(c) for c in rows[header_idx]]
    def col(*names):
        hl = [h.lower() for h in header]
        for n in names:
            try: return hl.index(n.lower())
            except ValueError: pass
        return None

    ci_acct = col("Account Name")
    ci_bank = col("Bank", "Bank name")
    ci_ano  = col("Account No")
    ci_own  = col("Owning Entity")
    ci_dir  = col("Debit/Credit","Debit / Credit")
    ci_feat = col("Rule Feature","Rule feature")
    ci_text = col("Text")
    ci_head = col("Head","Outcome head","Outcome Head")
    ci_sub  = col("Sub-head","Outcome Sub-head","Sub-head 1","Subhead")

    rules = []
    for row in rows[header_idx+1:]:
        if all(c is None or _s(c)=='' for c in row): continue
        def g(idx): return _s(row[idx]) if idx is not None and idx < len(row) else ""
        acct = g(ci_acct); head = g(ci_head)
        if not acct or not head: continue
        rules.append({
            "account_name":  acct,
            "bank":          g(ci_bank),
            "account_no":    g(ci_ano),
            "owning_entity": g(ci_own),
            "direction":     g(ci_dir).lower(),
            "rule_feature":  g(ci_feat),
            "text":          g(ci_text),
            "head":          head,
            "subhead":       g(ci_sub),
        })
    print(f"  Classification Rules: {len(rules)} rules")
    return rules


# ── 2. Parse Claude Classification tab ────────────────────────────────────────

def parse_claude_classification(wb):
    ws  = wb["Claude Classification"]
    all_rows = [(i, [_s(c) for c in row]) for i, row in enumerate(ws.iter_rows(values_only=True))]

    account_master = {}
    pattern_rules  = []
    valid_heads    = {}
    gotchas        = []

    section = None
    for i, vals in all_rows:
        joined = " | ".join(v for v in vals if v)
        if not joined: continue

        # Detect section headers
        if "SECTION 1" in vals[0].upper() or "ACCOUNT MASTER" in vals[0].upper():
            section = "account"; continue
        if "SECTION 2" in vals[0].upper() or "NARRATION PATTERN" in vals[0].upper():
            section = "pattern"; continue
        if "SECTION 3" in vals[0].upper() or "HEAD / SUB-HEAD MASTER" in vals[0].upper():
            section = "heads"; continue
        if "SECTION 4" in vals[0].upper() or "SPECIAL RULES" in vals[0].upper() or "GOTCHA" in vals[0].upper():
            section = "gotcha"; continue

        first = vals[0] if vals else ""

        # ── Account Master ──
        if section == "account":
            # Row: Account Name | Account Type | Credit Head | Credit Sub | Debit Head | Debit Sub | Notes
            if first.startswith("SECTION") or first in ("Account Name","") : continue
            if len(vals) >= 5 and first:
                acct      = first
                cr_head   = vals[2] if len(vals) > 2 else ""
                cr_sub    = vals[3] if len(vals) > 3 else ""
                db_head   = vals[4] if len(vals) > 4 else ""
                db_sub    = vals[5] if len(vals) > 5 else ""
                acct_type = vals[1] if len(vals) > 1 else ""
                notes     = vals[6] if len(vals) > 6 else ""
                if acct and (cr_head or db_head or acct_type):
                    account_master[acct] = {
                        "account_type": acct_type,
                        "credit": {"head": cr_head, "subhead": cr_sub},
                        "debit":  {"head": db_head, "subhead": db_sub},
                        "notes":  notes,
                    }

        # ── Pattern Rules ──
        elif section == "pattern":
            # Row: Priority | Narration Contains | Debit/Credit | Account Filter | Head | Sub-head | Notes
            if first in ("Priority","") or first.startswith("SECTION"): continue
            if len(vals) >= 5 and first:
                priority   = first
                narration  = vals[1] if len(vals) > 1 else ""
                direction  = vals[2] if len(vals) > 2 else ""
                acct_filter= vals[3] if len(vals) > 3 else ""
                head       = vals[4] if len(vals) > 4 else ""
                subhead    = vals[5] if len(vals) > 5 else ""
                notes      = vals[6] if len(vals) > 6 else ""
                if narration and head:
                    pattern_rules.append({
                        "priority":      priority,
                        "narration":     narration,
                        "direction":     direction.lower() if direction else "any",
                        "account_filter":acct_filter,
                        "head":          head,
                        "subhead":       subhead,
                        "notes":         notes,
                    })

        # ── Head / Sub-head Master ──
        elif section == "heads":
            if first in ("Head","") or first.startswith("SECTION"): continue
            if first and len(vals) >= 2:
                head   = first
                subhd  = vals[1] if len(vals) > 1 else ""
                direction = vals[2] if len(vals) > 2 else ""
                desc   = vals[3] if len(vals) > 3 else ""
                if head not in valid_heads:
                    valid_heads[head] = {"subheads": [], "description": desc}
                if subhd and subhd not in valid_heads[head]["subheads"]:
                    valid_heads[head]["subheads"].append(subhd)

        # ── Gotchas ──
        elif section == "gotcha":
            if first in ("Rule #","") or first.startswith("SECTION"): continue
            if first and len(vals) >= 2:
                gotchas.append({
                    "rule": first,
                    "description": vals[1] if len(vals) > 1 else "",
                })

    print(f"  Account Master: {len(account_master)} accounts")
    print(f"  Pattern Rules:  {len(pattern_rules)} rules")
    print(f"  Valid Heads:    {len(valid_heads)} heads")
    print(f"  Gotcha Rules:   {len(gotchas)} rules")
    return account_master, pattern_rules, valid_heads, gotchas


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Opening: {INPUT}")
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    os.makedirs(_DATA, exist_ok=True)

    # Classification Rules
    print("Parsing Classification Rules tab...")
    rules = parse_classification_rules(wb)
    with open(os.path.join(_DATA,"classification_rules.json"),"w",encoding="utf-8") as f:
        json.dump(rules, f, indent=2, ensure_ascii=True)

    # Claude Classification
    print("\nParsing Claude Classification tab...")
    acct_master, pat_rules, valid_heads, gotchas = parse_claude_classification(wb)

    # Save account master
    with open(os.path.join(_DATA,"account_master.json"),"w",encoding="utf-8") as f:
        json.dump(acct_master, f, indent=2, ensure_ascii=True)

    # Save valid heads
    heads_out = {
        "heads": list(valid_heads.keys()),
        "subheads": {h: v["subheads"] for h, v in valid_heads.items()},
        "descriptions": {h: v["description"] for h, v in valid_heads.items()},
    }
    with open(os.path.join(_DATA,"valid_heads.json"),"w",encoding="utf-8") as f:
        json.dump(heads_out, f, indent=2, ensure_ascii=True)

    # Save raw pattern rules (for AI context)
    with open(os.path.join(_DATA,"pattern_rules_raw.json"),"w",encoding="utf-8") as f:
        json.dump(pat_rules, f, indent=2, ensure_ascii=True)

    # Save gotcha rules
    with open(os.path.join(_DATA,"gotcha_rules.json"),"w",encoding="utf-8") as f:
        json.dump(gotchas, f, indent=2, ensure_ascii=True)

    wb.close()
    print(f"\nAll data files updated in: {_DATA}")
    print("Done.")

if __name__ == "__main__":
    main()
