"""
rule_parser.py
Parses the "Classification Rules" tab from the CICO Excel file into classification_rules.json.
Run directly:  python rule_parser.py --input CICO__claude.xlsx [--sheet "Classification Rules"]
Or call parse_rules_from_excel(path, sheet_name) programmatically.
"""

import argparse
import json
import os
import sys

import openpyxl


RULES_JSON_PATH = os.path.join(os.path.dirname(__file__), "data", "classification_rules.json")

EXPECTED_HEADERS = [
    "Account Name",
    "Bank",
    "Account No",
    "Owning Entity",
    "Debit/Credit",
    "Check In",
    "Rule Feature",
    "Text",
    "Head",
    "Sub-head",
]


def _normalise(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_rules_from_excel(excel_path: str, sheet_name: str = "Classification Rules") -> list:
    """
    Parse classification rules from an Excel sheet.
    Returns a list of rule dicts ordered exactly as they appear in the sheet.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Cannot open workbook '{excel_path}': {exc}")

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {available}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet is empty.")

    # Detect header row (first row that contains "Account Name")
    header_row_idx = None
    for i, row in enumerate(rows):
        row_vals = [_normalise(c) for c in row]
        if "Account Name" in row_vals:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find header row with 'Account Name' in the sheet.")

    header = [_normalise(c) for c in rows[header_row_idx]]

    def col(*names):
        """Find first matching column name (case-insensitive)."""
        header_lower = [h.lower() for h in header]
        for name in names:
            try:
                return header_lower.index(name.lower())
            except ValueError:
                continue
        return None

    idx_account_name  = col("Account Name")
    idx_bank          = col("Bank", "Bank name")
    idx_account_no    = col("Account No")
    idx_owning_entity = col("Owning Entity")
    idx_direction     = col("Debit/Credit", "Debit / Credit")
    idx_check_in      = col("Check In", "Check in")
    idx_rule_feature  = col("Rule Feature", "Rule feature")
    idx_text          = col("Text")
    idx_head          = col("Head", "Outcome head", "Outcome Head")
    idx_subhead       = col("Sub-head", "Outcome Sub-head", "Sub-head 1", "Subhead")

    rules = []
    for row in rows[header_row_idx + 1 :]:
        # Skip completely empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        def g(idx):
            if idx is None or idx >= len(row):
                return ""
            return _normalise(row[idx])

        account_name  = g(idx_account_name)
        bank          = g(idx_bank)
        account_no    = g(idx_account_no)
        owning_entity = g(idx_owning_entity)
        direction     = g(idx_direction)
        check_in      = g(idx_check_in)      # usually "Narration"
        rule_feature  = g(idx_rule_feature)  # All | Contains | Starts with
        text          = g(idx_text)
        head          = g(idx_head)
        subhead       = g(idx_subhead)

        if not account_name or not head:
            continue  # skip incomplete rows

        rule = {
            "account_name":  account_name,
            "bank":          bank,
            "account_no":    account_no,
            "owning_entity": owning_entity,
            "direction":     direction.lower() if direction else "",
            "check_in":      check_in,
            "rule_feature":  rule_feature,
            "text":          text,
            "head":          head,
            "subhead":       subhead,
        }
        rules.append(rule)

    wb.close()
    return rules


def save_rules(rules: list, output_path: str = RULES_JSON_PATH):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(rules, f, indent=2, ensure_ascii=True)
    return len(rules)


def load_rules(path: str = RULES_JSON_PATH) -> list:
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser(description="Parse Classification Rules tab into JSON.")
    parser.add_argument("--input", required=True, help="Path to CICO Excel file")
    parser.add_argument(
        "--sheet",
        default="Classification Rules",
        help="Sheet name (default: 'Classification Rules')",
    )
    parser.add_argument(
        "--output",
        default=RULES_JSON_PATH,
        help=f"Output JSON path (default: {RULES_JSON_PATH})",
    )
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing '{args.sheet}' from '{args.input}' ...")
    try:
        rules = parse_rules_from_excel(args.input, args.sheet)
        count = save_rules(rules, args.output)
        print(f"Saved {count} rules -> {args.output}")
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
