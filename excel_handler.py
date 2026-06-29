"""
excel_handler.py
openpyxl-based Excel I/O that preserves formatting.
"""

import logging
import os
from datetime import datetime
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column indices (1-based) for output columns
COL_HEAD_IDX    = 15   # O
COL_SUBHEAD_IDX = 16   # P
COL_SOURCE_IDX  = 17   # Q

# Expected column headers in the bank dump
EXPECTED_COLS = [
    "Bank", "Account No", "Account Name", "Account No", "Owning Entity",
    "Operating Entity", "Value Date", "Transaction Type", "Transaction",
    "CCY", "Debit", "Credit", "Net", "Narration",
    "Head", "Sub-head 1", "Sub-head 2",
]


def list_dump_sheets(file_path: str) -> list:
    """Return sheet names that look like bank dump sheets."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    candidates = []
    for name in wb.sheetnames:
        n = name.lower()
        if "dump" in n or "bank stt" in n or "statement" in n:
            candidates.append(name)
    wb.close()
    return candidates


def load_bank_dump(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """
    Load the bank dump Excel file into a DataFrame.
    Handles encoding issues in narration text.
    If sheet_name is None, auto-detects the first dump sheet.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name is None:
            dumps = list_dump_sheets(file_path)
            sheet_name = dumps[0] if dumps else wb.sheetnames[0]
        ws = wb[sheet_name]

        rows = []
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cleaned = []
            for cell in row:
                if cell is None:
                    cleaned.append("")
                elif isinstance(cell, str):
                    try:
                        cleaned.append(cell.encode("utf-8", errors="replace").decode("utf-8"))
                    except Exception:
                        cleaned.append(str(cell))
                else:
                    cleaned.append(cell)

            if header is None:
                # Find the header row: must contain "Account Name" or "Narration"
                vals = [str(c).strip() for c in cleaned]
                if "Account Name" in vals or "Narration" in vals:
                    header = cleaned
            else:
                rows.append(cleaned)

        wb.close()

        if header is None:
            raise ValueError("Could not find header row in the Excel file.")

        # Normalise header length
        max_len = max(len(header), max((len(r) for r in rows), default=0))
        header  = (header + [""] * max_len)[:max_len]
        rows    = [(r + [""] * max_len)[:max_len] for r in rows]

        df = pd.DataFrame(rows, columns=header)

        # Remove completely empty rows
        df = df.dropna(how="all")
        df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)]
        df = df.reset_index(drop=True)

        return df

    except Exception as exc:
        raise RuntimeError(f"Failed to load bank dump '{file_path}': {exc}")


def validate_columns(df: pd.DataFrame) -> dict:
    """
    Check that all expected columns are present.
    Returns {'ok': bool, 'missing': list, 'extra': list}.
    """
    actual  = set(df.columns)
    # Only check the first 14 core columns (not Head/Sub-head which may be empty)
    required = set(EXPECTED_COLS[:14])
    missing  = required - actual
    # Remove duplicate "Account No" from check (it appears twice in spec)
    missing.discard("Account No")
    return {
        "ok":      len(missing) == 0,
        "missing": sorted(missing),
        "extra":   sorted(actual - set(EXPECTED_COLS)),
    }


def write_classified_output(
    source_path: str,
    df: pd.DataFrame,
    summary_data: dict,
    output_folder: Optional[str] = None,
) -> str:
    """
    Write classified data back to a new Excel file, preserving all formatting.
    Adds a Classification_Summary sheet.
    Returns the output file path.
    """
    # Determine output path
    base       = os.path.basename(source_path)
    name, ext  = os.path.splitext(base)
    date_str   = datetime.now().strftime("%Y%m%d")
    out_name   = f"{name}_CLASSIFIED_{date_str}{ext}"
    folder     = output_folder or os.path.dirname(source_path)
    out_path   = os.path.join(folder, out_name)

    # Load original workbook (preserves all formatting)
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active  # first sheet = bank dump

    # Find header row index in the workbook
    header_row_idx = _find_header_row(ws)

    # Build a mapping from col name → column letter in the workbook
    col_map = {}
    if header_row_idx:
        for cell in ws[header_row_idx]:
            val = str(cell.value or "").strip()
            if val:
                col_map[val] = cell.column

    # Map DataFrame index → workbook row number (header_row_idx + 1 + df_idx)
    data_start_row = (header_row_idx or 1) + 1

    head_col    = col_map.get("Head",       COL_HEAD_IDX)
    subhead_col = col_map.get("Sub-head 1", COL_SUBHEAD_IDX)
    source_col  = col_map.get("Sub-head 2", COL_SOURCE_IDX)

    for df_idx, row in df.iterrows():
        wb_row = data_start_row + df_idx
        ws.cell(row=wb_row, column=head_col).value    = row.get("Head", "")
        ws.cell(row=wb_row, column=subhead_col).value = row.get("Sub-head 1", "")
        ws.cell(row=wb_row, column=source_col).value  = row.get("Sub-head 2", "")

    # Add Classification_Summary sheet
    if "Classification_Summary" in wb.sheetnames:
        del wb["Classification_Summary"]
    _write_summary_sheet(wb, summary_data)

    wb.save(out_path)
    logger.info(f"Saved classified output → {out_path}")
    return out_path


def _find_header_row(ws) -> Optional[int]:
    """Return the 1-based row number of the header row."""
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        vals = [str(c or "").strip() for c in row]
        if "Account Name" in vals or "Narration" in vals:
            return i
    return 1


def _write_summary_sheet(wb, summary: dict):
    ws = wb.create_sheet("Classification_Summary")

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill    = PatternFill("solid", fgColor="D6E4F0")
    bold_font   = Font(bold=True)

    def hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    def sub(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = sub_fill
        c.font = bold_font

    row = 1
    hdr(ws, row, 1, "CICO Classification Summary")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Processing stats
    sub(ws, row, 1, "Processing Statistics")
    row += 1
    stats = [
        ("Total rows processed (ex. DSA)",   summary.get("total", 0)),
        ("DSA rows skipped",                  summary.get("dsa_skip", 0)),
        ("Classified by Layer 1 / Layer 2 (RULE)", summary.get("layer1", 0) + summary.get("layer2", 0)),
        ("Classified by Layer 3 (AI)",        summary.get("ai", 0)),
        ("Still unclassified (TBD)",          summary.get("tbd", 0)),
    ]
    for label, val in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1
    row += 1

    # Head-wise summary
    sub(ws, row, 1, "Head-wise Summary")
    row += 1
    heads_hdr = ["Head", "Row Count", "Sum Credits (₹L)", "Sum Debits (₹L)", "Net (₹L)"]
    for c, h in enumerate(heads_hdr, 1):
        hdr(ws, row, c, h)
    row += 1
    for entry in summary.get("head_summary", []):
        ws.cell(row=row, column=1, value=entry.get("head", ""))
        ws.cell(row=row, column=2, value=entry.get("count", 0))
        ws.cell(row=row, column=3, value=round(entry.get("credits", 0) / 1e5, 2))
        ws.cell(row=row, column=4, value=round(entry.get("debits", 0) / 1e5, 2))
        ws.cell(row=row, column=5, value=round(entry.get("net", 0) / 1e5, 2))
        row += 1
    row += 1

    # Gotcha fixes
    sub(ws, row, 1, "Gotcha Fixes Applied")
    row += 1
    fixes = summary.get("gotcha_fixes", [])
    if fixes:
        for f in fixes[:200]:   # cap at 200 rows for readability
            ws.cell(row=row, column=1, value=f.get("gotcha", ""))
            ws.cell(row=row, column=2, value=f.get("account", ""))
            ws.cell(row=row, column=3, value=str(f.get("narration", ""))[:80])
            ws.cell(row=row, column=4, value=f.get("old_head", ""))
            ws.cell(row=row, column=5, value=f.get("new_head", ""))
            row += 1
    else:
        ws.cell(row=row, column=1, value="No fixes required.")
        row += 1
    row += 1

    # Contra matching
    sub(ws, row, 1, "Contra Matching Results")
    row += 1
    contra = summary.get("contra", {})
    ws.cell(row=row, column=1, value="Dates fully resolved")
    ws.cell(row=row, column=2, value=contra.get("dates_resolved", 0))
    row += 1
    ws.cell(row=row, column=1, value="Debit/credit pairs matched")
    ws.cell(row=row, column=2, value=contra.get("pairs_matched", 0))
    row += 1
    ws.cell(row=row, column=1, value="Unmatched entries remaining")
    ws.cell(row=row, column=2, value=contra.get("unmatched_count", 0))
    row += 1
    for u in contra.get("unmatched", [])[:50]:
        ws.cell(row=row, column=1, value=str(u.get("date", "")))
        ws.cell(row=row, column=2, value=u.get("account", ""))
        ws.cell(row=row, column=3, value=str(u.get("narration", ""))[:80])
        ws.cell(row=row, column=4, value=u.get("net", 0))
        row += 1
    row += 1

    # New narration patterns (TBD rows)
    sub(ws, row, 1, "Unclassified Narration Patterns (TBD — candidates for new rules)")
    row += 1
    for p in summary.get("tbd_narrations", [])[:100]:
        ws.cell(row=row, column=1, value=p.get("account", ""))
        ws.cell(row=row, column=2, value=str(p.get("narration", ""))[:120])
        ws.cell(row=row, column=3, value=p.get("direction", ""))
        ws.cell(row=row, column=4, value=p.get("amount", 0))
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
